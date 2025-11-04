"""
Report Generation Utilities
"""
from django.utils import timezone
from .models import Report
import io


def generate_report(report_type, title, file_format, parameters, branch_id, user):
    """
    Generate report based on type
    """
    from apps.branches.models import Branch
    
    branch = None
    if branch_id:
        branch = Branch.objects.get(id=branch_id)
    
    # Create report record
    report = Report.objects.create(
        title=title,
        report_type=report_type,
        file_format=file_format,
        parameters=parameters,
        branch=branch,
        created_by=user
    )
    
    # Generate file based on format
    if file_format == Report.ReportFormat.PDF:
        file_content = generate_pdf_report(report_type, parameters, branch)
        filename = f"{title}.pdf"
    elif file_format == Report.ReportFormat.EXCEL:
        file_content = generate_excel_report(report_type, parameters, branch)
        filename = f"{title}.xlsx"
    elif file_format == Report.ReportFormat.CSV:
        file_content = generate_csv_report(report_type, parameters, branch)
        filename = f"{title}.csv"
    else:
        file_content = generate_json_report(report_type, parameters, branch)
        filename = f"{title}.json"
    
    # Save file
    from django.core.files.base import ContentFile
    report.file.save(filename, ContentFile(file_content.getvalue()))
    report.file_size = report.file.size
    report.is_generated = True
    report.generated_at = timezone.now()
    report.save()
    
    return report


def generate_pdf_report(report_type, parameters, branch):
    """
    Generate PDF report
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(2*cm, height-2*cm, f"Report: {report_type}")
    
    # Add report content based on type
    y = height - 4*cm
    p.setFont("Helvetica", 12)
    
    if report_type == 'financial':
        p.drawString(2*cm, y, "Financial Report")
        # Add financial data
    elif report_type == 'enrollment':
        p.drawString(2*cm, y, "Enrollment Report")
        # Add enrollment data
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    return buffer


def generate_excel_report(report_type, parameters, branch):
    """
    Generate Excel report
    """
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = report_type
    
    # Add headers
    ws['A1'] = 'Report Type'
    ws['B1'] = report_type
    
    # Add data based on type
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_csv_report(report_type, parameters, branch):
    """
    Generate CSV report
    """
    import csv
    
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    
    # Write headers
    writer.writerow(['Report Type', 'Generated At'])
    writer.writerow([report_type, timezone.now()])
    
    # Write data based on type
    
    output = io.BytesIO(buffer.getvalue().encode('utf-8'))
    return output


def generate_json_report(report_type, parameters, branch):
    """
    Generate JSON report
    """
    import json
    
    data = {
        'report_type': report_type,
        'generated_at': str(timezone.now()),
        'parameters': parameters,
        'data': {}
    }
    
    # Add data based on type
    
    buffer = io.BytesIO(json.dumps(data, indent=2).encode('utf-8'))
    return buffer